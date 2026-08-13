"""Print the sheet inventory of a Supplementary Excel workbook."""

from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook", type=Path)
    ap.add_argument("--preview-rows", type=int, default=6)
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.workbook, read_only=True, data_only=True)
    print(f"workbook: {args.workbook}")
    print(f"n_sheets: {len(wb.sheetnames)}\n")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"=== {name!r}  rows={ws.max_row}  cols={ws.max_column}")
        for i, row in enumerate(ws.iter_rows(max_row=args.preview_rows, values_only=True)):
            cells = ["" if c is None else str(c)[:38] for c in row[:14]]
            print(f"   r{i}: {cells}")
        print()


if __name__ == "__main__":
    main()
